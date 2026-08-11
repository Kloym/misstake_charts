import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk
import os
import sys
import re
import unicodedata
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. БАЗОВЫЕ ФУНКЦИИ И ИНТЕРФЕЙС
# ==========================================

root = tk.Tk()
root.withdraw()

def get_file_path(title_text):
    file_path = filedialog.askopenfilename(
        parent=root,
        title=title_text,
        filetypes=[("CSV/Excel файлы", "*.csv *.xlsx *.xls *.XLSX *.XLS"), ("Все файлы", "*.*")]
    )
    return file_path

def load_and_clean_data(filepath, file_label):
    print(f"⏳ Загрузка {file_label} файла: {os.path.basename(filepath)}...")
    
    df = None
    if filepath.lower().endswith(('.xlsx', '.xls')):
        df = pd.read_excel(filepath, dtype=str)
    else:
        for enc in ['utf-8-sig', 'utf-8', 'cp1251']:
            for sep in [';', ',']:
                try:
                    temp_df = pd.read_csv(filepath, sep=sep, encoding=enc, dtype=str, skip_blank_lines=True)
                    if len(temp_df.columns) > 1:
                        df = temp_df
                        break
                except Exception:
                    continue
            if df is not None:
                break
                
    if df is None:
        raise ValueError(f"Не удалось корректно прочитать файл: {filepath}")

    print(f"⏳ Очистка и нормализация {file_label} файла...")
    
    if 'Версия' in df.columns:
        df = df.drop(columns=['Версия'])
    
    if len(df.columns) < 9:
        raise ValueError(f"В {file_label} файле меньше 9 столбцов (найдено {len(df.columns)}).")
        
    df = df.iloc[:, :9].copy()
    df.columns = df.columns.astype(str).str.replace(r'\xa0', ' ', regex=True).str.strip(' "')
    
    if df.columns.duplicated().any():
        duplicate_columns = df.columns[df.columns.duplicated()].tolist()
        raise ValueError(f"После очистки обнаружены одинаковые названия столбцов: {duplicate_columns}")
    
    df = df.fillna('')
    for col in df.columns:
        df[col] = df[col].astype(str).str.replace(r'\xa0', ' ', regex=True).str.strip(' "')

    bool_keywords = ["обязательность", "использован", "допустимость", "критерии", "архивност"]
    for col in df.columns:
        if any(kw in col.lower() for kw in bool_keywords):
            df[col] = df[col].replace({'0': 'Нет', '1': 'Да', '0.0': 'Нет', '1.0': 'Да'})
        elif "анестез" in col.lower():
            df[col] = df[col].replace({'0': 'Местная', '1': 'Общая', '0.0': 'Местная', '1.0': 'Общая'})
        
    df = df[df.astype(bool).any(axis=1)]

    initial_len = len(df)
    df = df.drop_duplicates()
    if len(df) < initial_len:
        print(f"   🧹 Автоматически удалено {initial_len - len(df)} полностью идентичных строк.")

    return df

# ==========================================
# 2. ФУНКЦИИ ОБОГАЩЕНИЯ НОВОГО MSCRIT
# ==========================================

def get_program_dir():
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

def normalize_key(value, strip_leading_zeros=False):
    if pd.isna(value):
        return ""
    
    value = unicodedata.normalize("NFKC", str(value))
    value = (
        value.replace("\xa0", " ")
             .replace("\u200b", "")
             .strip(" \"'")
             .upper()
    )
    
    if re.fullmatch(r"\d+\.0+", value):
        value = value.split(".", 1)[0]
        
    if strip_leading_zeros and value.isdigit():
        value = value.lstrip("0") or "0"
        
    return value

def clean_value(value):
    if pd.isna(value):
        return ""
    return str(value).replace("\xa0", " ").strip(' "')

def read_csv_safely(filepath, required_columns):
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        for separator in (";", ","):
            try:
                df = pd.read_csv(filepath, sep=separator, encoding=encoding, dtype=str, keep_default_na=False, skip_blank_lines=True)
                temp_cols = df.columns.astype(str).str.replace("\xa0", " ", regex=False).str.strip(' "')
                if all(col in temp_cols for col in required_columns):
                    df.columns = temp_cols
                    return df
            except Exception:
                continue
    raise ValueError(f"Не удалось прочитать файл '{filepath.name}' или отсутствуют колонки: {', '.join(required_columns)}")

def find_reference_file(prefix, program_dir):
    files = list(program_dir.glob(f"{prefix}*.csv"))
    if not files:
        raise FileNotFoundError(f"Не найден справочник '{prefix}*.csv' в папке:\n{program_dir}")
    if len(files) > 1:
        names = ", ".join(path.name for path in files)
        raise ValueError(f"❌ НАЙДЕНО НЕСКОЛЬКО ФАЙЛОВ ДЛЯ '{prefix}': {names}\nОставьте только ОДИН файл.")
    return files[0]

def build_mapping(prefix, key_columns, target_columns, program_dir, drop_pediatric=False, strip_zeros_cols=None):
    if isinstance(key_columns, str):
        key_columns = [key_columns]
        
    if strip_zeros_cols is None:
        strip_zeros_cols = []
        
    required_columns = key_columns + target_columns
    
    if drop_pediatric:
        required_columns.append("Профиль МП")

    filepath = find_reference_file(prefix, program_dir)
    print(f"⏳ Загрузка смежного справочника: {filepath.name}...")
    
    unique_required_cols = list(dict.fromkeys(required_columns))
    df = read_csv_safely(filepath, unique_required_cols)

    if drop_pediatric and "Профиль МП" in df.columns:
        mask = df["Профиль МП"].astype(str).str.lower().str.contains("детск", na=False)
        dropped_count = mask.sum()
        df = df[~mask].copy()
        if dropped_count > 0:
            print(f"   👶 Отфильтровано детских профилей: {dropped_count} шт.")

    for col in key_columns:
        should_strip = col in strip_zeros_cols
        df[col] = df[col].map(lambda v: normalize_key(v, strip_leading_zeros=should_strip))
        df = df[df[col] != ""].copy()
    
    for column in target_columns:
        df[column] = df[column].map(clean_value)
        
    return df[key_columns + target_columns].drop_duplicates()

def reorder_enriched_columns(df):
    """Сортировка колонок через словарь правил"""

    insert_rules = {
        "Код медицинской услуги": "Наименование",
        "Код хирургической операции": "Наименование операции"
    }
    
    # 2. Колонки, которые всегда идут в конец
    tail_cols = ["Профиль койки ФОМС V020", "Тариф"]
    
    # Собираем все добавленные колонки в одно множество для быстрой очистки
    added_cols = set(list(insert_rules.values()) + tail_cols)

    base_cols = [col for col in df.columns if col not in added_cols]
    
    final_cols = []
    
    # 4. Собираем новый список, автоматически применяя правила из словаря
    for col in base_cols:
        final_cols.append(col)
        
        target_col = insert_rules.get(col)
        if target_col and target_col in df.columns:
            final_cols.append(target_col)
            
    # 5. Добавляем хвост
    for col in tail_cols:
        if col in df.columns:
            final_cols.append(col)
            
    return df[final_cols]

def enrich_new_mscrit(df_new):
    print("\n⚙️ Обогащение нового справочника...")
    program_dir = get_program_dir()
    
    try:
        profms_df = build_mapping(
            prefix="profms", 
            key_columns=["Код МС или ВМП", "Код диагноза"], 
            target_columns=["МС или ВМП", "Профиль койки ФОМС V020"], 
            program_dir=program_dir,
            drop_pediatric=True,
            strip_zeros_cols=["Код МС или ВМП"]
        )
        profms_df = profms_df.rename(columns={'Код МС или ВМП': '_join_mes', 'Код диагноза': '_join_diag'})
        
        hopff_df = build_mapping("hopff", "Код опер.", ["Наимен.опер."], program_dir)
        hopff_df = hopff_df.rename(columns={'Код опер.': '_join_oper'})
        
        tarimu_df = build_mapping(
            prefix="tarimu", 
            key_columns="Код услуги", 
            target_columns=["Тариф"], 
            program_dir=program_dir,
            strip_zeros_cols=["Код услуги"]
        )
        tarimu_df = tarimu_df.rename(columns={'Код услуги': '_join_mes'})
        
    except Exception as e:
        print(f"\n{e}")
        print("⚠️ Обогащение прервано.")
        return None
    
    df_enriched = df_new.copy()
    initial_len = len(df_enriched)
    
    df_enriched['_join_mes'] = df_enriched["Код медицинской услуги"].map(lambda x: normalize_key(x, strip_leading_zeros=True))
    df_enriched['_join_oper'] = df_enriched["Код хирургической операции"].map(lambda x: normalize_key(x, strip_leading_zeros=False))
    df_enriched['_join_diag'] = df_enriched["Код диагноза"].map(lambda x: normalize_key(x, strip_leading_zeros=False))
    
    try:
        df_enriched = df_enriched.merge(profms_df, on=['_join_mes', '_join_diag'], how='left')
        df_enriched = df_enriched.merge(hopff_df, on=['_join_oper'], how='left', validate="m:1")
        df_enriched = df_enriched.merge(tarimu_df, on=['_join_mes'], how='left', validate="m:1")
    except pd.errors.MergeError as me:
        print(f"\n❌ ОШИБКА ДАННЫХ: Найдено недопустимое дублирование ключей в справочниках hopff или tarimu.")
        print(f"Техническая деталь: {me}")
        return None
    
    df_enriched = df_enriched.rename(columns={
        'МС или ВМП': 'Наименование',
        'Наимен.опер.': 'Наименование операции'
    })
    
    for col in ["Наименование", "Профиль койки ФОМС V020", "Наименование операции", "Тариф"]:
        df_enriched[col] = df_enriched[col].fillna("")
        
    df_enriched = df_enriched.drop(columns=['_join_mes', '_join_oper', '_join_diag'])
    df_enriched = reorder_enriched_columns(df_enriched)
    
    added_rows = len(df_enriched) - initial_len
    if added_rows > 0:
        print(f"   🔄 Из-за множественных профилей продублировано строк: {added_rows} шт.")
        
    print("✅ Обогащение успешно завершено!")
    return df_enriched

# ==========================================
# 3. GUI ДЛЯ РАЗРЕШЕНИЯ КОНФЛИКТОВ
# ==========================================

def resolve_duplicates_gui(df, file_label, key_cols):
    dupes_df = df[df.duplicated(subset=key_cols, keep=False)].copy()
    if dupes_df.empty: return df

    print(f"\n⚠️ ВНИМАНИЕ: В {file_label} файле обнаружены конфликтующие дубликаты! Запуск GUI...")
    window = tk.Toplevel(root)
    window.title(f"Разрешение конфликтов: {file_label} файл")
    window.geometry("1200x650")
    window.attributes('-topmost', True)
    window.grab_set()
    window.configure(bg="#F3F4F6")

    style = ttk.Style(window)
    style.theme_use("clam")
    style.configure("Treeview", background="#FFFFFF", foreground="#1F2937", rowheight=35,
                    fieldbackground="#FFFFFF", bordercolor="#E5E7EB", borderwidth=1, font=("Segoe UI", 10))
    style.map("Treeview", background=[("selected", "#DBEAFE")], foreground=[("selected", "#1E3A8A")])
    style.configure("Treeview.Heading", background="#F9FAFB", foreground="#374151",
                    font=("Segoe UI", 10, "bold"), borderwidth=1, relief="flat")
    style.map("Treeview.Heading", background=[("active", "#E5E7EB")])

    tk.Label(window, text=f"В {file_label} файле найдены конфликтующие дубликаты", 
             font=("Segoe UI", 14, "bold"), bg="#F3F4F6", fg="#111827").pack(pady=(20, 5))
    tk.Label(window, text="Отметьте галочкой (☑) те строки, которые нужно УДАЛИТЬ.\nВАЖНО: Для каждой связки должна остаться ровно одна правильная строка!", 
             font=("Segoe UI", 10), bg="#F3F4F6", fg="#4B5563").pack(pady=(0, 15))

    tree_frame = tk.Frame(window, bg="#F3F4F6")
    tree_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

    columns = ['check'] + list(df.columns)
    tree = ttk.Treeview(tree_frame, columns=columns, show='headings', selectmode="none")
    
    scroll_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
    scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=tree.xview)
    tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
    scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
    scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
    tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    tree.heading('check', text='Удалить?')
    tree.column('check', width=80, anchor='center', stretch=tk.NO)
    for col in df.columns:
        tree.heading(col, text=col)
        tree.column(col, width=130, anchor='center')

    item_to_df_index = {}
    for number, (df_index, row) in enumerate(dupes_df.iterrows()):
        item_id = f"row_{number}"
        item_to_df_index[item_id] = df_index
        tree.insert('', 'end', iid=item_id, values=['☐'] + list(row))

    def toggle_check(event):
        region = tree.identify("region", event.x, event.y)
        if region == "cell":
            column = tree.identify_column(event.x)
            if column == '#1': 
                item = tree.identify_row(event.y)
                if item:
                    vals = list(tree.item(item, "values"))
                    vals[0] = '☑' if vals[0] == '☐' else '☐'
                    tree.item(item, values=vals)

    tree.bind('<ButtonRelease-1>', toggle_check)
    result_state = {'action': 'stop', 'indices_to_drop': []}

    def on_continue():
        to_drop = [item_to_df_index[item_id] for item_id in tree.get_children() if tree.item(item_id, "values")[0] == '☑']
        result_state['action'] = 'continue'
        result_state['indices_to_drop'] = to_drop
        window.destroy()

    def on_stop():
        result_state['action'] = 'stop'
        window.destroy()

    btn_frame = tk.Frame(window, bg="#F3F4F6")
    btn_frame.pack(fill=tk.X, pady=20)
    
    btn_continue = tk.Button(btn_frame, text="✅ Удалить выделенные и Продолжить", command=on_continue, 
                             bg="#10B981", fg="white", activebackground="#059669", activeforeground="white",
                             font=("Segoe UI", 11, "bold"), relief="flat", padx=20, pady=10, cursor="hand2")
    btn_continue.pack(side=tk.LEFT, padx=30)
    
    btn_stop = tk.Button(btn_frame, text="❌ Прервать программу", command=on_stop, 
                         bg="#EF4444", fg="white", activebackground="#DC2626", activeforeground="white",
                         font=("Segoe UI", 11, "bold"), relief="flat", padx=20, pady=10, cursor="hand2")
    btn_stop.pack(side=tk.RIGHT, padx=30)

    window.wait_window()

    if result_state['action'] == 'stop':
        print("\n❌ Работа прервана пользователем на этапе разрешения дубликатов.")
        return None

    df_cleaned = df.drop(index=result_state['indices_to_drop'])
    conflict_keys = set(dupes_df[key_cols].itertuples(index=False, name=None))
    remaining_counts = df_cleaned.groupby(key_cols, dropna=False).size().to_dict()
    invalid_keys = [key for key in conflict_keys if remaining_counts.get(key, 0) != 1]
    
    if invalid_keys:
        print("\n❌ ОШИБКА: Вы удалили слишком много или слишком мало строк!")
        print("Для каждой конфликтной связки необходимо оставить РОВНО ОДНУ строку.")
        return None

    print(f"✅ Успешно удалено конфликтующих строк: {len(result_state['indices_to_drop'])} шт.")
    return df_cleaned

# ==========================================
# 4. ФОРМАТИРОВАНИЕ EXCEL ЛИСТОВ
# ==========================================

def format_excel_sheet_fast(ws, df, sample_size=5000):
    sample = df.head(sample_size)
    for column_number, column_name in enumerate(df.columns, start=1):
        lengths = sample[column_name].fillna("").astype(str).str.len()
        data_width = int(lengths.max()) if not lengths.empty else 0
        width = min(max(len(str(column_name)), data_width) + 3, 50)
        ws.column_dimensions[get_column_letter(column_number)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

def format_excel_sheet(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[col_letter].width = min(max_length + 3, 50)
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ==========================================
# 5. ГЛАВНАЯ ЛОГИКА
# ==========================================

def main():
    print("="*65)
    print(" 📊 Сравнение и обогащение справочников")
    print("="*65)
    
    print("\n[ШАГ 1] Выберите СТАРУЮ версию файла...")
    old_file = get_file_path("Выберите СТАРУЮ версию справочника")
    if not old_file: return

    print("\n[ШАГ 2] Выберите НОВУЮ версию файла...")
    new_file = get_file_path("Выберите НОВУЮ версию справочника")
    if not new_file: return
    
    if os.path.abspath(old_file) == os.path.abspath(new_file):
        print("\n❌ КРИТИЧЕСКАЯ ОШИБКА: Выбран один и тот же файл для старой и новой версии.")
        input("Нажмите Enter...")
        return

    df_old = load_and_clean_data(old_file, "СТАРОГО")
    df_new = load_and_clean_data(new_file, "НОВОГО")

    old_cols = list(df_old.columns)
    new_cols = list(df_new.columns)
    
    if old_cols != new_cols:
        print("\n❌ КРИТИЧЕСКАЯ ОШИБКА: Структура столбцов не совпадает!")
        input("Нажмите Enter...")
        return
        
    columns_list = old_cols

    try:
        key_cols = [columns_list[0], columns_list[1], columns_list[6]]
    except IndexError:
        print("❌ Ошибка: Недостаточно столбцов для извлечения ключей (A, B, G).")
        input("Нажмите Enter...")
        return
        
    for k in key_cols:
        df_old[k] = df_old[k].str.upper()
        df_new[k] = df_new[k].str.upper()

    old_before = len(df_old)
    new_before = len(df_new)
    df_old = df_old.drop_duplicates()
    df_new = df_new.drop_duplicates()
    if old_before - len(df_old) > 0:
        print(f"   🧹 После нормализации регистров удалено скрытых дублей в старом: {old_before - len(df_old)}")
    if new_before - len(df_new) > 0:
        print(f"   🧹 После нормализации регистров удалено скрытых дублей в новом: {new_before - len(df_new)}")

    for label, df in [("старом", df_old), ("новом", df_new)]:
        empty_key_mask = df[key_cols].eq('').any(axis=1)
        if empty_key_mask.any():
            print(f"\n❌ КРИТИЧЕСКАЯ ОШИБКА: В {label} файле найдено строк с пустыми ключами: {empty_key_mask.sum()} шт.")
            input("Нажмите Enter...")
            return

    print("\n🔍 Анализ на конфликты...")

    df_old = resolve_duplicates_gui(df_old, "СТАРОМ", key_cols)
    if df_old is None: 
        input("Нажмите Enter для выхода...")
        return
        
    df_new = resolve_duplicates_gui(df_new, "НОВОМ", key_cols)
    if df_new is None: 
        input("Нажмите Enter для выхода...")
        return

    # >>> ИНТЕГРАЦИЯ ОБОГАЩЕНИЯ <<<
    df_new_enriched = enrich_new_mscrit(df_new)
    
    if df_new_enriched is None:
        input("\nНажмите Enter для выхода...")
        return

    print("\n🚀 Формирование отличий...")
    
    df_old_idx = df_old.set_index(key_cols)
    df_new_idx = df_new.set_index(key_cols)

    added_index = df_new_idx.index.difference(df_old_idx.index, sort=False)
    removed_index = df_old_idx.index.difference(df_new_idx.index, sort=False)
    common_index = df_old_idx.index.intersection(df_new_idx.index, sort=False)
    
    output_data = [] 

    if len(removed_index):
        removed_rows = df_old_idx.loc[removed_index].reset_index()
        for _, row in removed_rows.iterrows():
            row_values = [row[col] for col in columns_list]
            output_data.append((['❌ Удалено'] + row_values, 'removed', []))

    if len(added_index):
        added_rows = df_new_idx.loc[added_index].reset_index()
        for _, row in added_rows.iterrows():
            row_values = [row[col] for col in columns_list]
            output_data.append((['✅ Добавлено'] + row_values, 'added', []))
            
    changes_count = 0
    total_changed_cells_count = 0
    
    if len(common_index):
        old_common = df_old_idx.reindex(common_index)
        new_common = df_new_idx.reindex(common_index)

        compared_cols = [col for col in columns_list if col not in key_cols]
        diff_mask = old_common[compared_cols].ne(new_common[compared_cols])
        changed_mask = diff_mask.any(axis=1)
        changed_indices = common_index[changed_mask.to_numpy()]
        
        changes_count = len(changed_indices)
        
        for idx in changed_indices:
            row_old = old_common.loc[idx]
            row_new = new_common.loc[idx]
            key_dict = dict(zip(key_cols, idx))
            out_row = ['⚠️ Изменено']
            changed_cols = []

            for excel_col_index, col in enumerate(columns_list, start=1):
                if col in key_cols:
                    out_row.append(key_dict[col])
                else:
                    val_old = row_old[col]
                    val_new = row_new[col]
                    if val_old != val_new:
                        out_row.append(f"{val_old} ➔ {val_new}")
                        changed_cols.append(excel_col_index)
                        total_changed_cells_count += 1
                    else:
                        out_row.append(val_new)
            
            output_data.append((out_row, 'modified', changed_cols))

    print(f"\n✅ Анализ завершен!")
    print(f"   - Удалено строк: {len(removed_index)} шт.")
    print(f"   - Добавлено строк: {len(added_index)} шт.")
    print(f"   - Изменено строк: {changes_count} шт.")
    print(f"   - Изменено ячеек (всего): {total_changed_cells_count} шт.")

    output_dir = os.path.dirname(new_file)
    output_filename = os.path.join(output_dir, 'Diff_Mscrit.xlsx')
    print(f"\n🎨 Формирование Excel-отчета...")
    
    wb = Workbook()
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    # --- ЛИСТ 1: ОБОГАЩЕННЫЙ СПРАВОЧНИК ---
    ws_enriched = wb.active
    ws_enriched.title = "Обогащенный справочник"
    
    # ПРАВКА 2: Очистка МЭС, начинающихся на '1' (кроме '183010')
    if "Код медицинской услуги" in df_new_enriched.columns:
        mes_codes = df_new_enriched["Код медицинской услуги"].map(
            lambda value: normalize_key(value, strip_leading_zeros=True)
        )
        remove_mask = mes_codes.str.startswith("1") & mes_codes.ne("183010")
        removed_count = int(remove_mask.sum())
        df_new_enriched = df_new_enriched.loc[~remove_mask].copy()
        print(f"   🧹 Удалено МЭС, начинающихся с '1': {removed_count} шт.")

    # ПРАВКА 3: Удаление столбца "Минимальное количество операций" с первого листа
    cols_to_drop = [c for c in df_new_enriched.columns if "минимальное количество" in c.lower()]
    df_new_enriched = df_new_enriched.drop(columns=cols_to_drop, errors='ignore')

    enriched_cols = list(df_new_enriched.columns)
    ws_enriched.append(enriched_cols)
    
    for cell in ws_enriched[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin

    for row in df_new_enriched.itertuples(index=False, name=None):
        ws_enriched.append(row)
            
    format_excel_sheet_fast(ws_enriched, df_new_enriched)
    
    ws_diff = wb.create_sheet(title="Сравнение версий")

    # --- ЛИСТ 2: СРАВНЕНИЕ ---
    if output_data:
        fill_added = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_removed = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_modified_row = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fill_modified_cell = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        header = ['Статус'] + columns_list
        ws_diff.append(header)
        
        for cell in ws_diff[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_thin

        for row_idx_offset, (row_data, status, changed_cols) in enumerate(output_data):
            ws_diff.append(row_data)
            current_row = ws_diff[row_idx_offset + 2] 
            row_fill = fill_added if status == 'added' else fill_removed if status == 'removed' else fill_modified_row

            for cell_idx, cell in enumerate(current_row):
                cell.border = border_thin
                cell.alignment = Alignment(vertical="center")
                if status == 'modified' and cell_idx in changed_cols:
                    cell.fill = fill_modified_cell
                    cell.font = Font(bold=True, color="9C0006")
                else:
                    cell.fill = row_fill

        format_excel_sheet(ws_diff)
    else:
        ws_diff.append(["Файлы абсолютно идентичны. Отчет не требуется."])

    try:
        wb.save(output_filename)
        print(f"\n🚀 ГОТОВО! Файл сохранен: {output_filename}")
    except PermissionError:
        print(f"\n❌ ОШИБКА ДОСТУПА: Не удалось сохранить файл {output_filename}. Возможно, он открыт в Excel.")
        
    input("\nНажмите Enter, чтобы выйти...")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ Произошла непредвиденная ошибка: {e}")
        input("Нажмите Enter для выхода...")